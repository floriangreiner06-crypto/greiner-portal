#!/usr/bin/env python3
"""
Stundensatz-Kalkulation Template Generator
==========================================
TAG 169: Erstellt Template aus Excel-Vorlage mit echten BWA-Daten

1. Liest Excel-Vorlage: docs/2021-10-11_Stundenverrechnungssatz berechnen(1).xls
2. Befüllt mit validen Werten aus BWA (loco_journal_accountings)
3. Erstellt Template für Planungstool
"""

import sys
import os
from datetime import datetime, date
from pathlib import Path

# Projekt-Pfad
sys.path.insert(0, str(Path(__file__).parent.parent))

from api.db_connection import get_db
from api.db_utils import db_session, locosoft_session, row_to_dict
from api.db_connection import convert_placeholders

# Import Filter-Funktion
try:
    from api.controlling_api import build_firma_standort_filter
except ImportError:
    # Fallback: Einfache Filter-Funktion
    def build_firma_standort_filter(firma, standort):
        """Einfache Filter-Funktion für Werkstatt"""
        if standort == 1:
            # Deggendorf: Stellantis (branch=1, subsidiary=1) + Hyundai (branch=2, subsidiary=2)
            umsatz = "AND ((branch_number = 1 AND subsidiary_to_company_ref = 1) OR (branch_number = 2 AND subsidiary_to_company_ref = 2))"
            einsatz = "AND ((SUBSTRING(CAST(nominal_account_number AS TEXT), 6, 1) = '1' AND subsidiary_to_company_ref = 1) OR (SUBSTRING(CAST(nominal_account_number AS TEXT), 6, 1) = '2' AND subsidiary_to_company_ref = 2))"
            kosten = einsatz
        elif standort == 2:
            # Hyundai: subsidiary=2
            umsatz = "AND subsidiary_to_company_ref = 2"
            einsatz = "AND subsidiary_to_company_ref = 2"
            kosten = einsatz
        elif standort == 3:
            # Landau: branch=3, subsidiary=1, Konto-Endziffer=2
            umsatz = "AND branch_number = 3 AND subsidiary_to_company_ref = 1"
            einsatz = "AND SUBSTRING(CAST(nominal_account_number AS TEXT), 6, 1) = '2' AND subsidiary_to_company_ref = 1"
            kosten = einsatz
        else:
            umsatz = einsatz = kosten = ""
        return umsatz, einsatz, kosten, f"Standort {standort}"

# =============================================================================
# EXCEL-STRUKTUR ANALYSIEREN
# =============================================================================

def analyse_excel_struktur(excel_path: str):
    """
    Analysiert die Excel-Vorlage und gibt Struktur zurück.
    """
    # Prüfe Dateiformat
    is_xls = str(excel_path).lower().endswith('.xls')
    
    if is_xls:
        # Altes .xls Format - verwende xlrd
        try:
            import xlrd
            # xlrd 2.0+ unterstützt nur .xlsx, für .xls brauchen wir xlrd < 2.0
            # Versuche zuerst mit xlrd
            try:
                wb = xlrd.open_workbook(excel_path)
            except:
                # Fallback: Versuche mit pandas
                import pandas as pd
                df = pd.read_excel(excel_path, sheet_name=None, engine='xlrd')
                # Konvertiere zu xlrd-ähnlichem Format
                class FakeWB:
                    def __init__(self, sheets):
                        self.sheet_names_ = list(sheets.keys())
                    def sheet_names(self):
                        return self.sheet_names_
                wb = FakeWB(df)
                structure = {'sheets': {}}
                for name, sheet_df in df.items():
                    structure['sheets'][name] = {'rows': []}
                    for row_idx, row in sheet_df.iterrows():
                        row_data = []
                        for col_idx, val in enumerate(row):
                            if pd.notna(val):
                                row_data.append({
                                    'col': col_idx,
                                    'col_letter': chr(65 + col_idx) if col_idx < 26 else 'AA',
                                    'row': row_idx + 2,  # +2 wegen Header
                                    'value': val
                                })
                        if row_data:
                            structure['sheets'][name]['rows'].append(row_data)
                return structure, wb
            
            structure = {
                'sheets': {}
            }
            
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                structure['sheets'][sheet_name] = {
                    'rows': []
                }
                
                # Alle Zeilen analysieren
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.value:
                            row_data.append({
                                'col': col_idx,
                                'col_letter': xlrd.colname(col_idx),
                                'row': row_idx + 1,
                                'value': cell.value,
                                'type': xlrd.biffh.XL_CELL_TEXT if cell.ctype == xlrd.XL_CELL_TEXT else 'number'
                            })
                    if row_data:
                        structure['sheets'][sheet_name]['rows'].append(row_data)
            
            return structure, wb
            
        except ImportError:
            print("ERROR: xlrd nicht installiert!")
            print("Bitte installieren: pip install xlrd")
            return None, None
    else:
        # Neues .xlsx Format - verwende openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            structure = {
                'sheets': {}
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                structure['sheets'][sheet_name] = {
                    'rows': []
                }
                
                # Erste 100 Zeilen analysieren
                for row_idx, row in enumerate(ws.iter_rows(max_row=100, values_only=False), 1):
                    row_data = []
                    for cell in row:
                        if cell.value is not None:
                            row_data.append({
                                'col': cell.column,
                                'col_letter': cell.column_letter,
                                'row': row_idx,
                                'value': cell.value
                            })
                    if row_data:
                        structure['sheets'][sheet_name]['rows'].append(row_data)
            
            return structure, wb
            
        except ImportError:
            print("ERROR: openpyxl nicht installiert!")
            print("Bitte installieren: pip install openpyxl")
            return None, None

# =============================================================================
# BWA-DATEN FÜR WERKSTATT LADEN
# =============================================================================

def lade_bwa_werkstatt_daten(standort: int = 1, jahr: int = None, monat: int = None):
    """
    Lädt BWA-Daten für Werkstatt aus loco_journal_accountings.
    
    Werkstatt-Konten:
    - Umsatz: 840000-849999
    - Einsatz: 740000-749999
    - Variable Kosten: 415xxx, 435xxx, 455xxx-456xxx, 487xxx, 491xxx-497xxx (mit KST 3)
    - Direkte Kosten: 4xxxxx mit KST 3 (ohne Variable)
    """
    if jahr is None:
        jahr = date.today().year
    if monat is None:
        monat = date.today().month
    
    # Geschäftsjahr bestimmen
    if monat >= 9:
        gj_start_jahr = jahr
    else:
        gj_start_jahr = jahr - 1
    
    # Letztes vollständiges Geschäftsjahr
    vj_gj_start = gj_start_jahr - 1
    vj_von = f"{vj_gj_start}-09-01"
    vj_bis = f"{vj_gj_start + 1}-09-01"
    
    result = {
        'standort': standort,
        'geschaeftsjahr': f"{vj_gj_start}/{str(vj_gj_start + 1)[2:]}",
        'von': vj_von,
        'bis': vj_bis,
        'umsatz': 0,
        'einsatz': 0,
        'db1': 0,
        'variable_kosten': 0,
        'direkte_kosten': 0,
        'db2': 0,
        'stunden_verkauft': 0,
        'stundensatz': 0,
        'kosten_pro_stunde': 0
    }
    
    # Standort-Filter (verwende build_firma_standort_filter)
    firma_filter_umsatz, firma_filter_einsatz, firma_filter_kosten, standort_name = build_firma_standort_filter('1', str(standort))
    
    try:
        # BWA-Daten sind in unserer PostgreSQL-DB gespiegelt (nicht Locosoft!)
        with db_session() as conn:
            cursor = conn.cursor()
            
            # 1. Umsatz (840000-849999) - aus gespiegelter loco_journal_accountings
            cursor.execute(f"""
                SELECT COALESCE(SUM(
                    CASE WHEN debit_or_credit='H' THEN posted_value ELSE -posted_value END
                )/100.0, 0) as umsatz
                FROM loco_journal_accountings
                WHERE accounting_date >= %s AND accounting_date < %s
                  AND nominal_account_number BETWEEN 840000 AND 849999
                  {firma_filter_umsatz}
            """, (vj_von, vj_bis))
            row = cursor.fetchone()
            result['umsatz'] = float(row[0] or 0) if row else 0
            
            # 2. Einsatz (740000-749999)
            cursor.execute(f"""
                SELECT COALESCE(SUM(
                    CASE WHEN debit_or_credit='S' THEN posted_value ELSE -posted_value END
                )/100.0, 0) as einsatz
                FROM loco_journal_accountings
                WHERE accounting_date >= %s AND accounting_date < %s
                  AND nominal_account_number BETWEEN 740000 AND 749999
                  {firma_filter_einsatz}
            """, (vj_von, vj_bis))
            row = cursor.fetchone()
            result['einsatz'] = float(row[0] or 0) if row else 0
            
            # 3. DB1 = Umsatz - Einsatz
            result['db1'] = result['umsatz'] - result['einsatz']
            
            # 4. Variable Kosten (415xxx, 435xxx, 455xxx-456xxx, 487xxx, 491xxx-497xxx mit KST 3)
            # KST 3 = 5. Ziffer = 3
            variable_konten = [
                (415000, 415999),  # Variable Kosten Gruppe 1
                (435000, 435999),  # Variable Kosten Gruppe 2
                (455000, 456999),  # Variable Kosten Gruppe 3
                (487000, 487999),  # Variable Kosten Gruppe 4
                (491000, 497999),  # Variable Kosten Gruppe 5
            ]
            
            variable_kosten_sum = 0
            for von, bis in variable_konten:
                cursor.execute(f"""
                    SELECT COALESCE(SUM(
                        CASE WHEN debit_or_credit='S' THEN posted_value ELSE -posted_value END
                    )/100.0, 0) as kosten
                    FROM loco_journal_accountings
                    WHERE accounting_date >= %s AND accounting_date < %s
                      AND nominal_account_number BETWEEN %s AND %s
                      AND CAST(SUBSTRING(CAST(nominal_account_number AS TEXT), 5, 1) AS INTEGER) = 3
                      {firma_filter_kosten}
                """, (vj_von, vj_bis, von, bis))
                row = cursor.fetchone()
                variable_kosten_sum += float(row[0] or 0) if row else 0
            
            result['variable_kosten'] = variable_kosten_sum
            
            # 5. Direkte Kosten (4xxxxx mit KST 3, ohne Variable)
            # Alle Kosten-Konten (400000-499999) mit KST 3, aber ohne Variable Kosten
            cursor.execute(f"""
                SELECT COALESCE(SUM(
                    CASE WHEN debit_or_credit='S' THEN posted_value ELSE -posted_value END
                )/100.0, 0) as kosten
                FROM loco_journal_accountings
                WHERE accounting_date >= %s AND accounting_date < %s
                  AND nominal_account_number BETWEEN 400000 AND 499999
                  AND CAST(SUBSTRING(CAST(nominal_account_number AS TEXT), 5, 1) AS INTEGER) = 3
                  AND NOT (
                      (nominal_account_number BETWEEN 415000 AND 415999) OR
                      (nominal_account_number BETWEEN 435000 AND 435999) OR
                      (nominal_account_number BETWEEN 455000 AND 456999) OR
                      (nominal_account_number BETWEEN 487000 AND 487999) OR
                      (nominal_account_number BETWEEN 491000 AND 497999)
                  )
                  {firma_filter_kosten}
            """, (vj_von, vj_bis))
            row = cursor.fetchone()
            result['direkte_kosten'] = float(row[0] or 0) if row else 0
            
            # 6. DB2 = DB1 - Variable Kosten - Direkte Kosten
            result['db2'] = result['db1'] - result['variable_kosten'] - result['direkte_kosten']
            
            # 7. Stunden verkauft (aus labours, AW zu Stunden) - aus Locosoft!
            with locosoft_session() as conn_loco:
                cursor_loco = conn_loco.cursor()
                cursor_loco.execute(f"""
                    SELECT COALESCE(SUM(l.time_units), 0) as aw_verkauft
                    FROM labours l
                    JOIN invoices i ON l.invoice_number = i.invoice_number 
                        AND l.invoice_type = i.invoice_type
                    JOIN orders o ON l.order_number = o.number
                    WHERE i.invoice_date >= %s AND i.invoice_date < %s
                      AND l.is_invoiced = true
                      AND i.is_canceled = false
                      AND o.subsidiary = {standort if standort in [1, 2] else 1}
                """, (vj_von, vj_bis))
                row = cursor_loco.fetchone()
            aw_verkauft = float(row[0] or 0) if row else 0
            result['stunden_verkauft'] = aw_verkauft / 6.0  # AW zu Stunden
            
            # 8. Stundensatz = Umsatz / Stunden
            if result['stunden_verkauft'] > 0:
                result['stundensatz'] = result['umsatz'] / result['stunden_verkauft']
            
            # 9. Kosten pro Stunde = (Variable + Direkte Kosten) / Stunden
            if result['stunden_verkauft'] > 0:
                result['kosten_pro_stunde'] = (result['variable_kosten'] + result['direkte_kosten']) / result['stunden_verkauft']
            
    except Exception as e:
        print(f"Fehler beim Laden der BWA-Daten: {e}")
        import traceback
        traceback.print_exc()
    
    return result

# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    # Excel-Datei finden (mit URL-Encoding im Namen)
    docs_dir = Path(__file__).parent.parent / 'docs'
    excel_files = list(docs_dir.glob('*Stundenverrechnung*.xls*'))
    if not excel_files:
        print(f"ERROR: Excel-Datei nicht gefunden in {docs_dir}")
        sys.exit(1)
    excel_path = excel_files[0]
    
    if not excel_path.exists():
        print(f"ERROR: Excel-Datei nicht gefunden: {excel_path}")
        sys.exit(1)
    
    print("=" * 80)
    print("STUNDENSATZ-KALKULATION TEMPLATE GENERATOR")
    print("=" * 80)
    print()
    
    # Excel-Struktur analysieren
    print("1. Analysiere Excel-Struktur...")
    structure, wb = analyse_excel_struktur(str(excel_path))
    
    if structure:
        print(f"   Gefundene Sheets: {list(structure['sheets'].keys())}")
        for sheet_name, sheet_data in structure['sheets'].items():
            print(f"   - {sheet_name}: {len(sheet_data['rows'])} Zeilen analysiert")
            print(f"\n   Detaillierte Struktur von '{sheet_name}':")
            # Erste 20 Zeilen ausgeben
            for row in sheet_data['rows'][:20]:
                print(f"     Zeile {row[0]['row']}: ", end="")
                for cell in row[:10]:  # Erste 10 Spalten
                    val_str = str(cell['value'])[:30]  # Max 30 Zeichen
                    print(f"{cell['col_letter']}:{val_str} | ", end="")
                print()
    
    print()
    
    # BWA-Daten laden
    print("2. Lade BWA-Daten für Werkstatt...")
    for standort in [1, 2, 3]:
        standort_name = {1: 'Deggendorf', 2: 'Hyundai DEG', 3: 'Landau'}[standort]
        print(f"   Standort {standort} ({standort_name}):")
        bwa_data = lade_bwa_werkstatt_daten(standort=standort)
        print(f"     Geschäftsjahr: {bwa_data['geschaeftsjahr']}")
        print(f"     Umsatz: {bwa_data['umsatz']:,.2f} €")
        print(f"     Einsatz: {bwa_data['einsatz']:,.2f} €")
        print(f"     DB1: {bwa_data['db1']:,.2f} €")
        print(f"     Variable Kosten: {bwa_data['variable_kosten']:,.2f} €")
        print(f"     Direkte Kosten: {bwa_data['direkte_kosten']:,.2f} €")
        print(f"     DB2: {bwa_data['db2']:,.2f} €")
        print(f"     Stunden verkauft: {bwa_data['stunden_verkauft']:,.2f} h")
        print(f"     Stundensatz: {bwa_data['stundensatz']:,.2f} €/h")
        print(f"     Kosten pro Stunde: {bwa_data['kosten_pro_stunde']:,.2f} €/h")
        print()
    
    print()
    print("=" * 80)
    print("TEMPLATE ERSTELLEN")
    print("=" * 80)
    
    # Template-Datei erstellen
    template_path = Path(__file__).parent.parent / 'docs' / 'STUNDENSATZ_KALKULATION_TEMPLATE_BEFUELLT.md'
    
    with open(template_path, 'w', encoding='utf-8') as f:
        f.write("# Stundensatz-Kalkulation Template (Befüllt mit BWA-Daten)\n")
        f.write("## TAG 169: Template für genauere Werkstatt-Planung\n\n")
        f.write("**Erstellt:** " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n\n")
        f.write("### BWA-Daten (Vorjahr 2024/25)\n\n")
        
        for standort in [1, 2, 3]:
            standort_name = {1: 'Deggendorf', 2: 'Hyundai DEG', 3: 'Landau'}[standort]
            bwa_data = lade_bwa_werkstatt_daten(standort=standort)
            
            f.write(f"#### Standort {standort}: {standort_name}\n\n")
            f.write("| Kennzahl | Wert |\n")
            f.write("|----------|------|\n")
            f.write(f"| **Umsatz (Jahr)** | {bwa_data['umsatz']:,.2f} € |\n")
            f.write(f"| **Einsatz (Jahr)** | {bwa_data['einsatz']:,.2f} € |\n")
            f.write(f"| **DB1 (Jahr)** | {bwa_data['db1']:,.2f} € |\n")
            f.write(f"| **Variable Kosten (Jahr)** | {bwa_data['variable_kosten']:,.2f} € |\n")
            f.write(f"| **Direkte Kosten (Jahr)** | {bwa_data['direkte_kosten']:,.2f} € |\n")
            f.write(f"| **DB2 (Jahr)** | {bwa_data['db2']:,.2f} € |\n")
            f.write(f"| **Stunden verkauft (Jahr)** | {bwa_data['stunden_verkauft']:,.2f} h |\n")
            f.write(f"| **Stundensatz** | {bwa_data['stundensatz']:,.2f} €/h |\n")
            f.write(f"| **Kosten pro Stunde** | {bwa_data['kosten_pro_stunde']:,.2f} €/h |\n")
            f.write(f"| **DB1 pro Stunde** | {(bwa_data['db1'] / bwa_data['stunden_verkauft'] if bwa_data['stunden_verkauft'] > 0 else 0):,.2f} €/h |\n")
            f.write(f"| **DB2 pro Stunde** | {(bwa_data['db2'] / bwa_data['stunden_verkauft'] if bwa_data['stunden_verkauft'] > 0 else 0):,.2f} €/h |\n")
            f.write("\n")
        
        f.write("### Excel-Struktur\n\n")
        if structure:
            for sheet_name, sheet_data in structure['sheets'].items():
                f.write(f"#### Sheet: {sheet_name}\n\n")
                f.write("| Zeile | Spalte | Wert |\n")
                f.write("|-------|--------|------|\n")
                for row in sheet_data['rows'][:30]:  # Erste 30 Zeilen
                    for cell in row[:8]:  # Erste 8 Spalten
                        val_str = str(cell['value']).replace('|', '\\|')[:50]
                        f.write(f"| {cell['row']} | {cell['col_letter']} | {val_str} |\n")
                f.write("\n")
    
    print(f"✅ Template erstellt: {template_path}")
    print()
    print("=" * 80)
    print("Nächste Schritte:")
    print("1. Excel-Struktur analysieren und Mapping erstellen")
    print("2. Template mit BWA-Daten befüllen")
    print("3. Template in Planungstool integrieren")
    print("=" * 80)

