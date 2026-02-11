#!/usr/bin/env python3
"""
Erstellt aus der VFW-CSV eine formatierte Excel-Datei (.xlsx) mit openpyxl.
Bearbeitbare Werte + Gesamtdurchschnitt (Formeln).
Benötigt: pip install openpyxl  (oder: python3 -m venv .venv && .venv/bin/pip install openpyxl)
Verwendung: python scripts/vfw_liste_to_excel.py [csv-pfad] [xlsx-pfad]
"""

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_CSV = ROOT / "docs" / "vfw_lohnsteuer_2023_2024.csv"
DEFAULT_XLSX = ROOT / "docs" / "vfw_lohnsteuer_2023_2024.xlsx"

SPALTEN = [
    ("standort_name", "Standort"),
    ("kennzeichen", "Kennzeichen"),
    ("vin", "FIN/VIN"),
    ("kom_nr", "Kom.Nr."),
    ("vfw_status", "Status"),
    ("marke", "Marke"),
    ("modell", "Modell"),
    ("ez", "EZ"),
    ("km", "km"),
    ("antriebsart", "Antriebsart"),
    ("upe_brutto", "UPE brutto (€)"),
    ("geldwert_vorteil_monat", "1 % / Monat (€)"),
    ("versteuerung_hinweis", "Versteuerung"),
    ("eingang", "Eingang"),
    ("abmeldung", "Abmeldung"),
    ("verkauft_am", "Verkauft am"),
]


def to_num(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def main():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Bitte zuerst openpyxl installieren:")
        print("  pip install openpyxl")
        print("  oder im Projekt: python3 -m venv .venv && .venv/bin/pip install openpyxl")
        print("  dann: .venv/bin/python scripts/vfw_liste_to_excel.py")
        sys.exit(1)

    csv_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_CSV
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_XLSX
    if not csv_path.exists():
        print(f"CSV nicht gefunden: {csv_path}")
        sys.exit(1)

    rows = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VFW Liste"

    col_keys = [k for k, _ in SPALTEN]
    headers = [label for _, label in SPALTEN]
    num_cols = len(headers)
    data_start = 2
    data_end = data_start + len(rows) - 1

    upe_col = geldwert_col = None
    for i, key in enumerate(col_keys):
        if key == "upe_brutto":
            upe_col = i + 1
        if key == "geldwert_vorteil_monat":
            geldwert_col = i + 1

    # Header
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Daten
    for r_idx, r in enumerate(rows, start=data_start):
        for c_idx, key in enumerate(col_keys, 1):
            val = r.get(key)
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            if key in ("upe_brutto", "geldwert_vorteil_monat"):
                n = to_num(val)
                cell.value = n if n is not None else None
                cell.number_format = "#,##0.00"
            elif key == "km":
                n = to_num(val)
                cell.value = int(n) if n is not None else None
            else:
                cell.value = (val or "").strip() or None

    # Spaltenbreiten
    for c, w in enumerate([14, 12, 8, 26, 10, 38, 12, 10, 18, 14, 14, 24, 12, 12, 12], 1):
        if c <= num_cols:
            ws.column_dimensions[get_column_letter(c)].width = min(w, 50)

    # Autofilter auf Kopfzeile + Daten (in Excel: Filter-Dropdowns zum Filtern)
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{data_end}"

    # Gesamtdurchschnitt
    summary_row = data_end + 2
    if upe_col:
        ws.cell(row=summary_row, column=1, value="Ø UPE brutto (€):").font = Font(bold=True)
        cell_upe = ws.cell(row=summary_row, column=upe_col)
        cell_upe.value = f"=AVERAGE({get_column_letter(upe_col)}{data_start}:{get_column_letter(upe_col)}{data_end})"
        cell_upe.number_format = "#,##0.00"
        cell_upe.font = Font(bold=True)
    if geldwert_col:
        ws.cell(row=summary_row + 1, column=1, value="Ø geldwerter Vorteil/Monat (€):").font = Font(bold=True)
        cell_g = ws.cell(row=summary_row + 1, column=geldwert_col)
        cell_g.value = f"=AVERAGE({get_column_letter(geldwert_col)}{data_start}:{get_column_letter(geldwert_col)}{data_end})"
        cell_g.number_format = "#,##0.00"
        cell_g.font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=1, value="Hinweis: UPE- und 1%-Werte in der Tabelle korrigieren/ergänzen – Durchschnitte aktualisieren sich automatisch.").font = Font(italic=True)
    ws.merge_cells(start_row=summary_row + 3, start_column=1, end_row=summary_row + 3, end_column=min(6, num_cols))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"Excel (.xlsx) erstellt: {xlsx_path} ({len(rows)} Zeilen, Durchschnitte in Zeile {summary_row})")


if __name__ == "__main__":
    main()
