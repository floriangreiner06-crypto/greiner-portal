#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export: Locosoft mitarbeiterrelevante Tabellen/Views als Excel
Für Vanessa Groll (1033) und Christian Aichinger (1005).
Ausgabe: Windows-Sync (z. B. für Vanessa/Christian zur Prüfung).
"""
import os
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Ziel: Windows-Sync (Server-Pfad)
OUTPUT_DIR = '/mnt/greiner-portal-sync'
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'Locosoft_Mitarbeiter_Uebersicht_Vanessa_Christian.xlsx')

# Locosoft employee_number für Vanessa Groll und Christian Aichinger
EMP_IDS = [1005, 1033]  # Christian Aichinger, Vanessa Groll

# Tabellen/Views mit Mitarbeiterbezug (Spalte für Filter)
TABLES_WITH_EMPLOYEE = [
    ('absence_calendar', 'employee_number'),
    ('employees', 'employee_number'),
    ('employees_history', 'employee_number'),
    ('employees_breaktimes', 'employee_number'),
    ('employees_group_mapping', 'employee_number'),
    ('employees_worktimes', 'employee_number'),
]
# appointments: mehrere employee-Spalten
TABLES_OTHER = [
    'absence_reasons',
    'absence_types',
    'year_calendar',
    'year_calendar_day_off_codes',
    'year_calendar_subsidiary_mapping',
]
# appointments: created_by_employee, bring_employee_no, return_employee_no
APPOINTMENTS_EMP_COLS = ['created_by_employee', 'bring_employee_no', 'return_employee_no', 'lock_by_employee', 'modified_by_employee', 'locked_by_employee']


def loco_conn():
    return psycopg2.connect(
        host='10.80.80.8',
        database='loco_auswertung_db',
        user='loco_auswertung_benutzer',
        password='loco'
    )


def get_columns(cursor, table_name):
    cursor.execute("""
        SELECT column_name, data_type
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
    """, (table_name,))
    return cursor.fetchall()


def fetch_data(cursor, table_name, emp_col, emp_ids):
    cols = [r[0] for r in get_columns(cursor, table_name)]
    placeholders = ','.join(['%s'] * len(emp_ids))
    q = f'SELECT * FROM {table_name} WHERE "{emp_col}" IN ({placeholders}) ORDER BY 1,2'
    cursor.execute(q, emp_ids)
    return cols, cursor.fetchall()


def fetch_appointments_for_employees(cursor, emp_ids):
    """Appointments wo einer der MA vorkommt (created_by, bring_employee_no, return_employee_no)."""
    cols = [r[0] for r in get_columns(cursor, 'appointments')]
    placeholders = ','.join(['%s'] * len(emp_ids))
    conditions = ' OR '.join([f'{c} IN ({placeholders})' for c in APPOINTMENTS_EMP_COLS])
    q = f'SELECT * FROM appointments WHERE {conditions} ORDER BY id LIMIT 500'
    params = emp_ids * len(APPOINTMENTS_EMP_COLS)
    cursor.execute(q, params)
    return cols, cursor.fetchall()


def fetch_table_sample(cursor, table_name, limit=100):
    cols = [r[0] for r in get_columns(cursor, table_name)]
    cursor.execute(f'SELECT * FROM {table_name} LIMIT {limit}')
    return cols, cursor.fetchall()


def sheet_safe_name(name):
    return name[:31].replace('/', '-').replace('\\', '-')


def format_cell(value):
    if value is None:
        return ''
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def main():
    wb = Workbook()
    wb.remove(wb.active)

    conn = loco_conn()
    cur = conn.cursor()

    # --- Sheet: Übersicht ---
    ws_over = wb.create_sheet('Übersicht', 0)
    ws_over.append(['Locosoft DB: mitarbeiterrelevante Tabellen/Views'])
    ws_over.append(['Stand: ' + datetime.now().isoformat()[:19]])
    ws_over.append(['Mitarbeiter: Christian Aichinger (1005), Vanessa Groll (1033)'])
    ws_over.append([])
    ws_over.append(['Tabelle/View', 'Spalten (Name, Typ)', 'Hinweis'])
    for row in ws_over.iter_rows(min_row=1, max_row=4):
        for c in row:
            c.font = Font(bold=True)
    col_info = get_columns(cur, 'employees_history')  # dummy to have a get_columns call
    tables_all = [t[0] for t in TABLES_WITH_EMPLOYEE] + ['appointments', 'appointments_text'] + TABLES_OTHER
    for t in tables_all:
        cols = get_columns(cur, t)
        col_str = '; '.join([f"{c[0]} ({c[1]})" for c in cols[:15]])
        if len(cols) > 15:
            col_str += f' ... +{len(cols)-15} weitere'
        note = 'Daten für MA 1005, 1033' if t in [x[0] for x in TABLES_WITH_EMPLOYEE] or t == 'appointments' else 'Referenz/Sample'
        ws_over.append([t, col_str, note])
    ws_over.column_dimensions['A'].width = 28
    ws_over.column_dimensions['B'].width = 80
    ws_over.column_dimensions['C'].width = 22

    # --- Sheet: Alle Spalten (Tabelle | Spalte | Typ) zum Suchen nach Urlaub/Anspruch
    ws_cols = wb.create_sheet('Alle_Spalten_alle_Tabellen', 1)
    ws_cols.append(['Tabelle/View', 'Spalte', 'Datentyp'])
    for row in ws_cols.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = Font(bold=True)
    for t in tables_all:
        cols = get_columns(cur, t)
        for c in cols:
            ws_cols.append([t, c[0], c[1]])
    ws_cols.column_dimensions['A'].width = 28
    ws_cols.column_dimensions['B'].width = 32
    ws_cols.column_dimensions['C'].width = 28

    # --- Sheets pro Tabelle (mit employee_number) ---
    for table_name, emp_col in TABLES_WITH_EMPLOYEE:
        try:
            cols, rows = fetch_data(cur, table_name, emp_col, EMP_IDS)
        except Exception as e:
            cols, rows = get_columns(cur, table_name), []
            rows = [(f'Fehler: {e}',)]
            cols = [('Fehler', 'text')]
            if isinstance(cols[0], tuple):
                cols = [c[0] for c in cols]
        ws = wb.create_sheet(sheet_safe_name(table_name))
        ws.append([c if isinstance(c, str) else c[0] for c in cols])
        for r in rows:
            ws.append([format_cell(v) for v in r])
        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14
    # --- appointments ---
    try:
        cols, rows = fetch_appointments_for_employees(cur, EMP_IDS)
    except Exception as e:
        cols = [c[0] for c in get_columns(cur, 'appointments')]
        rows = [(f'Fehler: {e}',)]
    ws = wb.create_sheet('appointments')
    ws.append(cols)
    for r in rows:
        ws.append([format_cell(v) for v in r])
    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # --- appointments_text (nur wenn appointment_id in gefundenen appointments) ---
    try:
        cur.execute('SELECT column_name FROM information_schema.columns WHERE table_schema=%s AND table_name=%s ORDER BY ordinal_position', ('public', 'appointments_text'))
        cols = [r[0] for r in cur.fetchall()]
        ws = wb.create_sheet('appointments_text')
        ws.append(cols)
        cur.execute('SELECT * FROM appointments_text LIMIT 50')
        for r in cur.fetchall():
            ws.append([format_cell(v) for v in r])
    except Exception as e:
        ws = wb.create_sheet('appointments_text')
        ws.append(['Fehler', str(e)])

    # --- Referenz-Tabellen (vollständig oder Sample) ---
    for table_name in TABLES_OTHER:
        try:
            cols, rows = fetch_table_sample(cur, table_name)
        except Exception as e:
            cols = ['Fehler']
            rows = [[str(e)]]
        ws = wb.create_sheet(sheet_safe_name(table_name))
        ws.append([c if isinstance(c, str) else c[0] for c in cols])
        for r in rows:
            ws.append([format_cell(v) for v in r])
        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18

    conn.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f'Excel erstellt: {OUTPUT_FILE}')
    print(f'  (Windows-Sync: \\\\Srvrdb01\\Allgemein\\Greiner Portal\\Greiner_Portal_NEU\\Server\\Locosoft_Mitarbeiter_Uebersicht_Vanessa_Christian.xlsx)')


if __name__ == '__main__':
    main()
