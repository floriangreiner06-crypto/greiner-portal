#!/usr/bin/env python3
"""
Bewertungskatalog (AUTO1) mit Fahrzeugbestand befüllen.
============================================================
Datenquelle: Locosoft (Standard) oder eAutoSeller (--eautoseller).
Kopiert die Blanko-Excel 1:1 und befüllt nur die Datenzeilen (ab Zeile 16)
genau nach den Vorgaben des Templates (Spalten Nr, FIN, Hersteller, …).
VK-Preis und Anzahl Schlüssel werden nicht exportiert (bleiben leer).

Verwendung:
  python scripts/export_bewertungskatalog.py --min-standzeit 300
  python scripts/export_bewertungskatalog.py --eautoseller --min-standzeit 300   # Quelle: eAutoSeller API
  python scripts/export_bewertungskatalog.py --standort 1

Workstream: verkauf
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import date, datetime

# Projekt-Root
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from api.db_utils import locosoft_session
from api.standort_utils import STANDORT_NAMEN, build_locosoft_filter_bestand
from psycopg2.extras import RealDictCursor

FAHRZEUGTYP_NAMEN = {
    "N": "Neuwagen",
    "G": "Gebrauchtwagen",
    "D": "Vorführwagen",
    "V": "Vermietwagen",
    "T": "Tauschfahrzeug",
}

# Template-Pfad (Sync = Windows sichtbar)
DEFAULT_TEMPLATE = "/mnt/greiner-portal-sync/docs/workstreams/verkauf/Bewertungskatalog  blanko.xlsx"
# Fallback wenn Sync nicht gemountet
FALLBACK_TEMPLATE = os.path.join(
    os.path.dirname(__file__), "..", "docs", "workstreams", "verkauf", "Bewertungskatalog  blanko.xlsx"
)


def fetch_locosoft_mwst_halter_by_vin(vins: list[str]) -> dict[str, dict]:
    """
    Holt aus Locosoft pro VIN: out_sale_type (MwSt), dealer_vehicle_type, Anzahl Halter.
    previous_owner_counter in Locosoft = Anzahl Vorbesitzer → „Anzahl Halter“ = Vorbesitzer + 1 (1. Halter, 2. Halter, …).
    Returns: { vin_upper: {"out_sale_type", "dealer_vehicle_type", "previous_owner_counter": 1|2|3…}, ... }
    """
    if not vins:
        return {}
    vins_clean = [str(v).strip().upper() for v in vins if v and str(v).strip()]
    if not vins_clean:
        return {}
    result = {}
    with locosoft_session() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            """
            SELECT v.vin, v.previous_owner_counter, dv.out_sale_type, dv.dealer_vehicle_type
            FROM vehicles v
            JOIN dealer_vehicles dv
                ON v.dealer_vehicle_number = dv.dealer_vehicle_number
                AND v.dealer_vehicle_type = dv.dealer_vehicle_type
            WHERE v.vin IS NOT NULL AND TRIM(v.vin) != ''
            AND UPPER(TRIM(v.vin)) = ANY(%s)
            """,
            (vins_clean,),
        )
        for row in cur.fetchall():
            vin = (row.get("vin") or "").strip().upper()
            if vin:
                # Anzahl Halter = 1. Halter, 2. Halter, … (previous_owner_counter = Vorbesitzeranzahl)
                cnt = row.get("previous_owner_counter")
                anzahl_halter = (int(cnt) + 1) if cnt is not None else 1
                result[vin] = {
                    "out_sale_type": row.get("out_sale_type"),
                    "dealer_vehicle_type": row.get("dealer_vehicle_type"),
                    "previous_owner_counter": anzahl_halter,
                }
    return result


def fetch_locosoft_verkaufte_vins(vins: list[str]) -> set[str]:
    """
    Liefert die Menge der VINs, die in Locosoft bereits verkauft sind (out_invoice_date IS NOT NULL).
    So können sie aus dem Bewertungskatalog-Export ausgeschlossen werden.
    """
    if not vins:
        return set()
    vins_clean = [str(v).strip().upper() for v in vins if v and str(v).strip()]
    if not vins_clean:
        return set()
    with locosoft_session() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT UPPER(TRIM(v.vin)) AS vin
            FROM vehicles v
            JOIN dealer_vehicles dv
                ON v.dealer_vehicle_number = dv.dealer_vehicle_number
                AND v.dealer_vehicle_type = dv.dealer_vehicle_type
            WHERE v.vin IS NOT NULL AND TRIM(v.vin) != ''
            AND dv.out_invoice_date IS NOT NULL
            AND UPPER(TRIM(v.vin)) = ANY(%s)
            """,
            (vins_clean,),
        )
        return {row[0] for row in cur.fetchall() if row[0]}


def get_bestand_fuer_bewertung(
    standort: int | None = None,
    inkl_vfw: bool = False,
    min_standzeit_tage: int | None = None,
    sort_nach_marke: bool = True,
) -> list[dict]:
    """Holt GW-Bestand (optional + VFW) aus Locosoft für Bewertungskatalog."""
    where_parts = [
        "dv.out_invoice_date IS NULL",
        "dv.dealer_vehicle_type IN ('G', 'D', 'T')" if inkl_vfw else "dv.dealer_vehicle_type = 'G'",
    ]
    params: list = []
    if standort:
        standort_filter = build_locosoft_filter_bestand(standort, nur_stellantis=False)
        if standort_filter:
            filter_sql = standort_filter.replace("AND ", "").replace("in_subsidiary", "dv.in_subsidiary")
            where_parts.append(filter_sql)
    if min_standzeit_tage is not None:
        where_parts.append(
            "(CURRENT_DATE - COALESCE(dv.in_arrival_date, dv.created_date)) > %s"
        )
        params.append(min_standzeit_tage)
    where_clause = " AND ".join(where_parts)

    order = (
        "ORDER BY COALESCE(mk.description, v.free_form_make_text), dv.dealer_vehicle_number"
        if sort_nach_marke
        else "ORDER BY dv.in_subsidiary, dv.dealer_vehicle_type, dv.dealer_vehicle_number"
    )

    query = f"""
        SELECT
            dv.dealer_vehicle_number,
            dv.dealer_vehicle_type,
            v.vin,
            v.license_plate,
            COALESCE(v.free_form_model_text, m.description) AS modell,
            v.first_registration_date AS ez,
            dv.out_sale_type,
            dv.out_sale_price,
            dv.in_subsidiary AS standort,
            dv.location AS lagerort,
            COALESCE(mk.description, v.free_form_make_text) AS hersteller,
            v.previous_owner_counter,
            v.accidents_counter,
            v.is_all_accidents_repaired,
            v.carkey_number
        FROM dealer_vehicles dv
        LEFT JOIN vehicles v
            ON dv.dealer_vehicle_number = v.dealer_vehicle_number
            AND dv.dealer_vehicle_type = v.dealer_vehicle_type
        LEFT JOIN models m
            ON dv.out_model_code = m.model_code AND dv.out_make_number = m.make_number
        LEFT JOIN makes mk ON dv.out_make_number = mk.make_number
        WHERE {where_clause}
        {order}
    """
    with locosoft_session() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(query, params)
        return [dict(row) for row in cur.fetchall()]


def get_bestand_aus_eautoseller(
    min_standzeit_tage: int | None = None,
    sort_nach_marke: bool = True,
) -> list[dict]:
    """
    Holt Fahrzeugbestand aus eAutoSeller (Swagger API, Fallback HTML).
    Liefert Zeilen im gleichen Format wie get_bestand_fuer_bewertung (vin, hersteller, modell, ez, …).
    """
    import json
    from lib.eautoseller_client import EAutosellerClient

    creds_path = os.path.join(os.path.dirname(__file__), "..", "config", "credentials.json")
    creds = {}
    if os.path.isfile(creds_path):
        try:
            with open(creds_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                creds = data.get("eautoseller", {})
        except Exception:
            pass
    # Swagger (api_key/client_secret) hat Vorrang; Fallback: username/password für HTML
    api_key = (creds.get("api_key") or creds.get("API_Key") or os.getenv("EAUTOSELLER_API_KEY") or "").strip()
    client_secret = (creds.get("client_secret") or creds.get("clientSecret") or os.getenv("EAUTOSELLER_CLIENT_SECRET") or "").strip()
    username = creds.get("username", os.getenv("EAUTOSELLER_USERNAME", "")) or "swagger"
    password = creds.get("password", os.getenv("EAUTOSELLER_PASSWORD", "")) or ""

    client = EAutosellerClient(
        username=username,
        password=password,
        loginbereich=creds.get("loginbereich", "kfz"),
    )

    # Zuerst Swagger API (nutzt api_key/client_secret aus credentials.json), Fallback HTML
    vehicles = []
    if api_key and client_secret:
        try:
            vehicles = client.get_vehicles_swagger(use_swagger=True)
        except Exception as e:
            print(f"eAutoSeller Swagger fehlgeschlagen: {e}", file=sys.stderr)
    if not vehicles and username and password and username != "swagger":
        try:
            vehicles = client.get_vehicle_list(active_only=True, fetch_hereinnahme=True)
        except Exception as e:
            print(f"eAutoSeller HTML-Fallback fehlgeschlagen: {e}", file=sys.stderr)
    if not vehicles and not api_key:
        raise RuntimeError("eAutoSeller-Credentials fehlen (config/credentials.json: eautoseller.api_key + client_secret oder username + password)")

    if not vehicles:
        return []

    # Standzeit berechnen wo fehlend
    today = date.today()
    for v in vehicles:
        if v.get("standzeit_tage") is None and v.get("hereinnahme"):
            try:
                if isinstance(v["hereinnahme"], str):
                    dt = datetime.strptime(v["hereinnahme"][:10], "%Y-%m-%d").date()
                else:
                    dt = v["hereinnahme"]
                v["standzeit_tage"] = (today - dt).days
            except Exception:
                pass

    # Filter: nur Standzeit > min_standzeit_tage
    if min_standzeit_tage is not None:
        vehicles = [v for v in vehicles if (v.get("standzeit_tage") or 0) > min_standzeit_tage]

    # Sortierung nach Marke
    if sort_nach_marke:
        vehicles = sorted(vehicles, key=lambda v: (v.get("marke") or "", v.get("vin") or ""))

    # Auf gleiches Zeilenformat wie Locosoft mappen (für befuellen())
    zeilen = []
    for v in vehicles:
        ez = None
        raw = v.get("_raw_swagger") or {}
        for key in ("firstRegistrationDate", "firstRegistration", "registrationDate", "year"):
            val = raw.get(key) or v.get("baujahr")
            if val is None:
                continue
            if isinstance(val, str) and len(val) >= 10:
                try:
                    ez = datetime.strptime(val[:10], "%Y-%m-%d").date()
                except Exception:
                    try:
                        ez = datetime.strptime(val[:10], "%d.%m.%Y").date()
                    except Exception:
                        pass
            elif isinstance(val, (int, float)) and 1990 < val < 2030:
                try:
                    ez = date(int(val), 1, 1)
                except Exception:
                    pass
            if ez is not None:
                break

        # VIN: mehrere API-Feldnamen prüfen (teilweise liefert eAutoSeller kein vin)
        vin = (
            (v.get("vin") or "")
            or (raw.get("vehicleIdentificationNumber") or "")
            or (raw.get("identificationNumber") or "")
            or (raw.get("vin") or "")
        )
        if isinstance(vin, str):
            vin = vin.strip()
        else:
            vin = str(vin).strip() if vin else ""

        kennzeichen = (
            raw.get("licensePlate")
            or raw.get("license_plate")
            or v.get("kennzeichen")
            or ""
        )
        if isinstance(kennzeichen, dict):
            kennzeichen = kennzeichen.get("number") or kennzeichen.get("value") or ""

        modell = (v.get("modell") or "").strip()
        # Vorlauf-Fahrzeuge ausschließen: oft ohne VIN oder Modell „Fahrgestell“/„Vorlauf“
        raw_modell = (raw.get("modelName") or raw.get("model") or modell) or ""
        if isinstance(raw_modell, dict):
            raw_modell = raw_modell.get("name", "") or str(raw_modell)
        is_vorlauf = (
            not vin
            or "fahrgestell" in modell.lower()
            or "vorlauf" in modell.lower()
            or "fahrgestell" in str(raw_modell).lower()
            or "vorlauf" in str(raw_modell).lower()
        )
        if is_vorlauf:
            continue

        # MwSt ausweisbar? (ja = Regelbesteuerung/F, nein = Differenz §25a/B) – Client oder raw
        out_sale_type = v.get("out_sale_type")
        if out_sale_type is None:
            for key in ("taxType", "saleType", "vatDeductible", "invoiceType", "mwst", "besteuerung"):
                val = raw.get(key)
                if val is None:
                    continue
                if isinstance(val, bool):
                    out_sale_type = "F" if val else "B"
                    break
                s = str(val).upper()
                if "REGEL" in s or "F" == s or "JA" in s or "TRUE" in s:
                    out_sale_type = "F"
                    break
                if "DIFF" in s or "25A" in s or "B" == s or "NEIN" in s or "FALSE" in s:
                    out_sale_type = "B"
                    break

        # Anzahl Halter – Client oder raw
        previous_owner_counter = v.get("previous_owner_counter")
        if previous_owner_counter is None:
            for key in ("previousOwnerCount", "ownerCount", "numberOfOwners", "vorbesitzer", "holderCount", "anzahlHalter"):
                val = raw.get(key)
                if val is not None:
                    try:
                        previous_owner_counter = int(val)
                        break
                    except (TypeError, ValueError):
                        pass

        zeilen.append({
            "vin": vin,
            "hersteller": v.get("marke") or "",
            "modell": modell,
            "ez": ez,
            "out_sale_type": out_sale_type,
            "previous_owner_counter": previous_owner_counter,
            "accidents_counter": None,
            "dealer_vehicle_type": "G",
            "license_plate": kennzeichen if isinstance(kennzeichen, str) else "",
            "standort": None,
            "standort_name": "eAutoSeller",
        })
    return zeilen


def format_ez(d: date | None) -> str:
    if d is None:
        return ""
    return d.strftime("%d.%m.%Y")


def mwst_ausweisbar(out_sale_type: str | None, dealer_vehicle_type: str | None = None) -> str:
    """
    MwSt ausweisbar? Entspricht kalkulation_helpers: dealer_vehicle_type hat Vorrang vor out_sale_type.
    G = Gebrauchtwagen → immer Regel (ja). D = Vorführwagen → Diff25a (nein). Dann out_sale_type F/B.
    """
    typ = (dealer_vehicle_type or "").strip().upper()
    # Wie in kalkulation_helpers: G → Regel (ja), D → Diff25a (nein)
    if typ == "G":
        return "ja"
    if typ == "D":
        return "nein"
    if out_sale_type is not None and str(out_sale_type).strip():
        return "nein" if str(out_sale_type).strip().upper() == "B" else "ja"
    return "ja"


def unfallfahrzeug(row: dict) -> str:
    acc = row.get("accidents_counter") or 0
    if acc > 0:
        return "ja"
    # is_all_accidents_repaired = false könnte auch „Unfall, repariert“ bedeuten
    return "nein" if acc == 0 else ""


def befuellen(template_path: str, ausgabe_path: str, zeilen: list[dict]) -> None:
    import openpyxl
    from openpyxl.styles import Font

    # 1. Blanko exakt kopieren (Layout, Formatierung, alle Sheets unverändert)
    shutil.copy2(template_path, ausgabe_path)
    # 2. Nur die Kopie öffnen und ausschließlich Datenzeilen befüllen (Feldbeschreibungen in Zeile 16 unberührt)
    wb = openpyxl.load_workbook(ausgabe_path)
    ws = wb["Template"]
    # Einheitliche Schrift für alle Datenzeilen (aus Vorlage Zeile 2 übernehmen, sonst Calibri 11)
    try:
        ref = ws.cell(2, 3).font
        ref_font = Font(name=(ref.name or "Calibri"), size=(ref.size or 11)) if ref else Font(name="Calibri", size=11)
    except Exception:
        ref_font = Font(name="Calibri", size=11)
    # Template: Zeile 16 = Feldbeschreibungen (Nr, FIN, Hersteller, …) – nicht überschreiben!
    # Ab Zeile 17 = Datenzeilen „Zu bewertende Fahrzeuge“
    start_row = 17
    for i, row in enumerate(zeilen, start=1):
        excel_row = start_row + i - 1
        for col, value in [
            (1, i),
            (2, row.get("vin") or ""),
            (3, row.get("hersteller") or ""),
            (4, row.get("modell") or ""),
            (5, format_ez(row.get("ez"))),
            (6, mwst_ausweisbar(row.get("out_sale_type"), row.get("dealer_vehicle_type"))),
            (7, row.get("previous_owner_counter") if row.get("previous_owner_counter") is not None else ""),
            (8, unfallfahrzeug(row)),
            (9, FAHRZEUGTYP_NAMEN.get(row.get("dealer_vehicle_type") or "", "")),
            (10, ""),
            (11, row.get("license_plate") or ""),
            (12, ""),
            (13, ""),
            (14, f"Standort: {row.get('standort_name') or STANDORT_NAMEN.get(row.get('standort'), 'Unbekannt')}"),
        ]:
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.font = ref_font
    wb.save(ausgabe_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Bewertungskatalog mit Fahrzeugbestand befüllen")
    parser.add_argument("--eautoseller", action="store_true", help="Datenquelle: eAutoSeller API (statt Locosoft)")
    parser.add_argument("--standort", type=int, default=None, help="1=DEG Opel, 2=HYU, 3=LAN (nur bei Locosoft)")
    parser.add_argument("--inkl-vfw", action="store_true", help="Vorführwagen und Tausch (nur bei Locosoft)")
    parser.add_argument("--min-standzeit", type=int, default=None, metavar="TAGE", help="Nur Fahrzeuge mit Standzeit über TAGE (eAutoSeller Standard: 250)")
    parser.add_argument("--no-sort-marke", action="store_true", help="Nicht nach Marke sortieren (Standard: nach Marke)")
    parser.add_argument("--template", default=None, help="Pfad zur Blanko-Excel")
    parser.add_argument("--out", default=None, help="Ausgabe-Pfad (Standard: Bewertungskatalog_befuellt_<datum>.xlsx)")
    args = parser.parse_args()

    template = args.template or (DEFAULT_TEMPLATE if os.path.isfile(DEFAULT_TEMPLATE) else FALLBACK_TEMPLATE)
    if not os.path.isfile(template):
        print(f"Template nicht gefunden: {template}", file=sys.stderr)
        print("Bitte 'Bewertungskatalog  blanko.xlsx' nach docs/workstreams/verkauf legen (oder Sync mounten).", file=sys.stderr)
        sys.exit(1)

    # Standzeit-Filter: eAutoSeller Standard 250 Tage, Locosoft nur wenn angegeben
    standzeit_filter = args.min_standzeit if args.min_standzeit is not None else (250 if args.eautoseller else None)

    if args.eautoseller:
        try:
            zeilen = get_bestand_aus_eautoseller(
                min_standzeit_tage=standzeit_filter,
                sort_nach_marke=not args.no_sort_marke,
            )
        except Exception as e:
            print(f"eAutoSeller-Fehler: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        zeilen = get_bestand_fuer_bewertung(
            standort=args.standort,
            inkl_vfw=args.inkl_vfw,
            min_standzeit_tage=args.min_standzeit,
            sort_nach_marke=not args.no_sort_marke,
        )

    if not zeilen:
        msg = "Keine Fahrzeuge (Quelle: " + ("eAutoSeller" if args.eautoseller else "Locosoft GW")
        if args.inkl_vfw and not args.eautoseller:
            msg += " + VFW/T"
        if standzeit_filter is not None:
            msg += f", Standzeit > {standzeit_filter} Tage"
        msg += ")."
        print(msg, file=sys.stderr)
        sys.exit(0)

    # eAutoSeller: bereits verkaufte VINs (Locosoft) aus Export entfernen
    if args.eautoseller and zeilen:
        vins_all = [str(z.get("vin") or "").strip() for z in zeilen if z.get("vin")]
        verkaufte = fetch_locosoft_verkaufte_vins(vins_all)
        if verkaufte:
            zeilen = [z for z in zeilen if str(z.get("vin") or "").strip().upper() not in verkaufte]
            if not zeilen:
                print("Keine Fahrzeuge nach Ausschluss bereits verkaufter (Locosoft).", file=sys.stderr)
                sys.exit(0)

    # eAutoSeller: MwSt ausweisbar + Anzahl Halter per VIN aus Locosoft nachziehen
    if args.eautoseller and zeilen:
        vins = [str(z.get("vin") or "").strip() for z in zeilen if z.get("vin")]
        loco_by_vin = fetch_locosoft_mwst_halter_by_vin(vins)
        for z in zeilen:
            vin_key = str(z.get("vin") or "").strip().upper()
            if vin_key and vin_key in loco_by_vin:
                data = loco_by_vin[vin_key]
                if data.get("out_sale_type") is not None:
                    z["out_sale_type"] = data["out_sale_type"]
                if data.get("dealer_vehicle_type") is not None:
                    z["dealer_vehicle_type"] = data["dealer_vehicle_type"]
                if data.get("previous_owner_counter") is not None:
                    z["previous_owner_counter"] = data["previous_owner_counter"]

    out_dir = os.path.dirname(template)
    if args.out:
        ausgabe = args.out
    else:
        suffix = "_eautoseller" if args.eautoseller else ""
        ausgabe = os.path.join(out_dir, f"Bewertungskatalog_befuellt{suffix}_{date.today().isoformat()}.xlsx")

    befuellen(template, ausgabe, zeilen)
    print(f"OK: {len(zeilen)} Fahrzeuge nach {ausgabe} geschrieben.")


if __name__ == "__main__":
    main()
