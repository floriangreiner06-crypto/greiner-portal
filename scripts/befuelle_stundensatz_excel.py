#!/usr/bin/env python3
"""
Stundensatz-Excel mit BWA-Daten befüllen
=========================================
TAG 169: Befüllt die Excel-Vorlage mit echten BWA-Daten

Mapping Excel → BWA:
- Löhne, Gehälter, Sozialabgaben → Personalkosten (410xxx, 420xxx) mit KST 3
- Lieferanten → Variable Kosten (415xxx, 435xxx, etc.) mit KST 3
- Mieten und Nebenkosten → Miete/Leasing (480xxx) mit KST 3
- Büro und Verwaltungskosten → Verwaltungskosten (4xxxxx mit KST 3)
- etc.
"""

import sys
from pathlib import Path
from datetime import datetime, date

sys.path.insert(0, str(Path(__file__).parent.parent))

from api.db_connection import get_db
from api.db_utils import db_session, locosoft_session, row_to_dict
from api.db_connection import convert_placeholders

try:
    from api.controlling_api import build_firma_standort_filter
except ImportError:
    def build_firma_standort_filter(firma, standort):
        if standort == 1:
            umsatz = "AND ((branch_number = 1 AND subsidiary_to_company_ref = 1) OR (branch_number = 2 AND subsidiary_to_company_ref = 2))"
            einsatz = "AND ((SUBSTRING(CAST(nominal_account_number AS TEXT), 6, 1) = '1' AND subsidiary_to_company_ref = 1) OR (SUBSTRING(CAST(nominal_account_number AS TEXT), 6, 1) = '2' AND subsidiary_to_company_ref = 2))"
            kosten = einsatz
        elif standort == 2:
            umsatz = "AND subsidiary_to_company_ref = 2"
            einsatz = "AND subsidiary_to_company_ref = 2"
            kosten = einsatz
        elif standort == 3:
            umsatz = "AND branch_number = 3 AND subsidiary_to_company_ref = 1"
            einsatz = "AND SUBSTRING(CAST(nominal_account_number AS TEXT), 6, 1) = '2' AND subsidiary_to_company_ref = 1"
            kosten = einsatz
        else:
            umsatz = einsatz = kosten = ""
        return umsatz, einsatz, kosten, f"Standort {standort}"

# =============================================================================
# KOSTEN-KATEGORIEN MAPPING (Excel → BWA-Konten)
# =============================================================================

KOSTEN_MAPPING = {
    'Löhne, Gehälter, Sozialabgaben': {
        'konten': [(410000, 429999)],  # Personalkosten
        'kst': 3,
        'beschreibung': 'Personalkosten Werkstatt (410xxx-429xxx)'
    },
    'Lieferanten': {
        'konten': [(415000, 415999), (435000, 435999)],  # Variable Kosten
        'kst': 3,
        'beschreibung': 'Variable Kosten (Lieferanten)'
    },
    'Mieten und Nebenkosten': {
        'konten': [(480000, 480999)],  # Miete/Leasing
        'kst': 3,
        'beschreibung': 'Miete/Leasing (480xxx)'
    },
    'Büro und Verwaltungskosten': {
        'konten': [(400000, 499999)],  # Alle Kosten mit KST 3 (ohne spezifische)
        'kst': 3,
        'beschreibung': 'Verwaltungskosten Werkstatt'
    },
    'Provisionen / Gratifikationen': {
        'konten': [(490000, 490999), (491000, 497999)],  # Provisionen
        'kst': 3,
        'beschreibung': 'Provisionen (490xxx-497xxx)'
    },
    'Training/Weiterbildung': {
        'konten': [(455000, 456999)],  # Variable Kosten Gruppe 3
        'kst': 3,
        'beschreibung': 'Training/Weiterbildung'
    },
    'Werkstattwagen': {
        'konten': [(450000, 450999)],  # Kfz-Kosten
        'kst': 3,
        'beschreibung': 'Kfz-Kosten Werkstatt (450xxx)'
    },
    'Werbung/Verkaufsförderung': {
        'konten': [(487000, 487999)],  # Variable Kosten Gruppe 4
        'kst': 3,
        'beschreibung': 'Werbung (487xxx)'
    },
    'Garantie': {
        'konten': [(455000, 456999)],  # Variable Kosten
        'kst': 3,
        'beschreibung': 'Garantie-Kosten'
    },
    'Kulanz': {
        'konten': [(455000, 456999)],  # Variable Kosten
        'kst': 3,
        'beschreibung': 'Kulanz-Kosten'
    },
    'Frachtkosten': {
        'konten': [(455000, 456999)],  # Variable Kosten
        'kst': 3,
        'beschreibung': 'Frachtkosten'
    },
    'Hilfs- und Betriebsstoffe': {
        'konten': [(460000, 469999)],  # Betriebsstoffe
        'kst': 3,
        'beschreibung': 'Betriebsstoffe (460xxx)'
    },
    'Werkzeuge/Kleinteile': {
        'konten': [(470000, 479999)],  # Reparaturen/Instandhaltung
        'kst': 3,
        'beschreibung': 'Werkzeuge/Reparaturen (470xxx)'
    },
    'Entsorgung/Recycling': {
        'konten': [(455000, 456999)],  # Variable Kosten
        'kst': 3,
        'beschreibung': 'Entsorgung'
    },
    'Instandhaltung': {
        'konten': [(470000, 479999)],  # Reparaturen
        'kst': 3,
        'beschreibung': 'Instandhaltung (470xxx)'
    },
    'Aufw. für Diagnosegeräte': {
        'konten': [(440000, 449999)],  # AfA/Investitionen
        'kst': 3,
        'beschreibung': 'AfA/Investitionen (440xxx)'
    }
}

# =============================================================================
# BWA-DATEN LADEN
# =============================================================================

def lade_kosten_kategorie(kategorie_name: str, standort: int, jahr: int = None, monat: int = None):
    """
    Lädt Kosten für eine spezifische Kategorie aus BWA.
    """
    if jahr is None:
        jahr = date.today().year
    if monat is None:
        monat = date.today().month
    
    # Geschäftsjahr bestimmen
    if monat >= 9:
        gj_start_jahr = jahr
    else:
        gj_start_jahr = jahr - 1
    
    vj_gj_start = gj_start_jahr - 1
    vj_von = f"{vj_gj_start}-09-01"
    vj_bis = f"{vj_gj_start + 1}-09-01"
    
    mapping = KOSTEN_MAPPING.get(kategorie_name, {})
    if not mapping:
        return 0.0
    
    konten = mapping.get('konten', [])
    kst = mapping.get('kst', 3)
    
    firma_filter_umsatz, firma_filter_einsatz, firma_filter_kosten, _ = build_firma_standort_filter('1', str(standort))
    
    try:
        with db_session() as conn:
            cursor = conn.cursor()
            
            kosten_sum = 0.0
            for von, bis in konten:
                cursor.execute(f"""
                    SELECT COALESCE(SUM(
                        CASE WHEN debit_or_credit='S' THEN posted_value ELSE -posted_value END
                    )/100.0, 0) as kosten
                    FROM loco_journal_accountings
                    WHERE accounting_date >= %s AND accounting_date < %s
                      AND nominal_account_number BETWEEN %s AND %s
                      AND CAST(SUBSTRING(CAST(nominal_account_number AS TEXT), 5, 1) AS INTEGER) = %s
                      {firma_filter_kosten}
                """, (vj_von, vj_bis, von, bis, kst))
                row = cursor.fetchone()
                kosten_sum += float(row[0] or 0) if row else 0
            
            return kosten_sum
            
    except Exception as e:
        print(f"Fehler beim Laden von {kategorie_name}: {e}")
        return 0.0

# =============================================================================
# EXCEL BEFÜLLEN
# =============================================================================

def befuelle_excel(excel_path: str, standort: int, output_path: str = None):
    """
    Befüllt Excel mit BWA-Daten für einen Standort.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        # Versuche .xlsx zu erstellen (konvertiere von .xls)
        if excel_path.endswith('.xls'):
            # Für .xls: Verwende pandas zum Konvertieren
            try:
                import pandas as pd
                df = pd.read_excel(excel_path, sheet_name=None, engine='xlrd')
                # Speichere als .xlsx
                temp_xlsx = excel_path.replace('.xls', '_temp.xlsx')
                with pd.ExcelWriter(temp_xlsx, engine='openpyxl') as writer:
                    for sheet_name, sheet_df in df.items():
                        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                excel_path = temp_xlsx
            except:
                print("WARNUNG: Kann .xls nicht konvertieren. Verwende Original.")
                return False
        
        wb = load_workbook(excel_path)
        
        # Finde Sheet "Externer Verrechnungssatz"
        sheet_name = None
        for name in wb.sheetnames:
            if 'Verrechnungssatz' in name or 'Stunden' in name:
                sheet_name = name
                break
        
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        
        # Mapping: Suche nach Kosten-Kategorien und befülle
        standort_name = {1: 'Deggendorf', 2: 'Hyundai DEG', 3: 'Landau'}[standort]
        
        print(f"\nBefülle Excel für {standort_name}...")
        
        # Durchsuche Zeilen nach Kosten-Kategorien
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Prüfe ob es eine bekannte Kategorie ist
                    for kategorie_name in KOSTEN_MAPPING.keys():
                        if kategorie_name.lower() in cell.value.lower():
                            # Finde Wert-Spalte (normalerweise Spalte C oder D)
                            wert_cell = ws.cell(row=row_idx, column=3)  # Spalte C
                            if wert_cell.value is None or wert_cell.value == 0:
                                # Lade Wert aus BWA
                                wert = lade_kosten_kategorie(kategorie_name, standort)
                                wert_cell.value = wert
                                print(f"  Zeile {row_idx}: {kategorie_name} = {wert:,.2f} €")
                            break
        
        # Speichere
        if output_path is None:
            output_path = excel_path.replace('.xls', f'_befuellt_{standort_name}.xlsx').replace('.xlsx', f'_befuellt_{standort_name}.xlsx')
        
        wb.save(output_path)
        print(f"\n✅ Excel gespeichert: {output_path}")
        return True
        
    except Exception as e:
        print(f"Fehler beim Befüllen der Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Befüllt Stundensatz-Excel mit BWA-Daten')
    parser.add_argument('--standort', type=int, choices=[1, 2, 3], default=1, help='Standort (1=DEG, 2=HYU, 3=LAN)')
    parser.add_argument('--excel', type=str, help='Pfad zur Excel-Datei')
    parser.add_argument('--output', type=str, help='Ausgabe-Pfad')
    
    args = parser.parse_args()
    
    if args.excel:
        excel_path = args.excel
    else:
        # Suche Excel im docs-Verzeichnis
        docs_dir = Path(__file__).parent.parent / 'docs'
        excel_files = list(docs_dir.glob('*Stundenverrechnung*.xls*'))
        if not excel_files:
            print(f"ERROR: Excel-Datei nicht gefunden in {docs_dir}")
            sys.exit(1)
        excel_path = str(excel_files[0])
    
    print("=" * 80)
    print("STUNDENSATZ-EXCEL BEFÜLLEN")
    print("=" * 80)
    print(f"Excel: {excel_path}")
    print(f"Standort: {args.standort}")
    print()
    
    success = befuelle_excel(excel_path, args.standort, args.output)
    
    if success:
        print("\n✅ Erfolgreich befüllt!")
    else:
        print("\n❌ Fehler beim Befüllen!")
        sys.exit(1)

